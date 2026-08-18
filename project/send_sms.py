from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime
import gc
from io import BytesIO
import logging
from pathlib import Path
import re
import shutil
import tempfile
import time
import tomllib
from typing import Any, Dict, Iterable, Iterator, Sequence

from openpyxl import load_workbook
import pandas as pd
import requests


logger = logging.getLogger(__name__)

DEFAULT_CONFIG_PATH = Path("config.toml")
DEFAULT_BATCH_SIZE = 10
DEFAULT_SEND_TIMEOUT = 10.0
DEFAULT_ADD_DELAY_SECONDS = 45.0
DEFAULT_FIND_DELAY_SECONDS = 20.0
SUPPLEMENT_SHEET_NAME = "Doplnit"
GATE_PHONE_COLUMN_INDEX = 1
SUPPLEMENT_PHONE_COLUMN_INDEX = 0
PHONE_PATTERN = re.compile(r"^\+\d{9,15}$")


class SmsError(Exception):
    """Base exception for the SMS application."""


class ConfigurationError(SmsError):
    """Raised when configuration is missing or invalid."""


class SpreadsheetError(SmsError):
    """Raised when spreadsheet input is unreadable or invalid."""


@dataclass(frozen=True)
class GateConfig:
    id: int
    phone: str
    password: str
    sheet: str


@dataclass(frozen=True)
class AppConfig:
    gateway_base: str
    excel_path: Path
    gates: Dict[int, GateConfig]

    @property
    def send_endpoint(self) -> str:
        return f"{self.gateway_base.rstrip('/')}/send-sms"

    def get_gate(self, gate_id: int) -> GateConfig:
        try:
            return self.gates[gate_id]
        except KeyError as exc:
            raise ConfigurationError(f"Gate {gate_id} is not configured.") from exc


@dataclass(frozen=True)
class SmsResult:
    ok: bool
    phone: str
    message: str
    status_code: int | None = None
    payload: Any | None = None
    error: str | None = None


@dataclass(frozen=True)
class SheetRow:
    row_number: int
    raw_value: str
    normalized: str | None
    status: str
    error: str | None = None


@dataclass(frozen=True)
class SheetAnalysis:
    source_name: str
    rows: list[SheetRow]
    valid_numbers: list[str]
    unique_numbers: list[str]
    duplicate_numbers: list[str]
    invalid_rows: list[SheetRow]
    blank_row_count: int

    @property
    def total_rows(self) -> int:
        return len(self.rows)

    @property
    def valid_row_count(self) -> int:
        return len(self.valid_numbers)

    @property
    def unique_row_count(self) -> int:
        return len(self.unique_numbers)

    @property
    def invalid_row_count(self) -> int:
        return len(self.invalid_rows)


@dataclass(frozen=True)
class SheetSaveResult:
    sheet_name: str
    rows_written: int
    numbers: list[str]
    backup_path: Path | None = None


def load_config(config_path: str | Path | None = None) -> AppConfig:
    path = Path(config_path) if config_path else DEFAULT_CONFIG_PATH

    try:
        raw_config = tomllib.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ConfigurationError(
            f"Configuration file '{path}' was not found. Create it from config.example.toml."
        ) from exc
    except tomllib.TOMLDecodeError as exc:
        raise ConfigurationError(f"Configuration file '{path}' is not valid TOML: {exc}") from exc

    gateway_base = raw_config.get("gateway_base", "").strip().rstrip("/")
    if not gateway_base:
        raise ConfigurationError("Missing 'gateway_base' in config.toml.")

    excel_path_raw = raw_config.get("excel_path", "").strip()
    if not excel_path_raw:
        raise ConfigurationError("Missing 'excel_path' in config.toml.")
    excel_path = Path(excel_path_raw)

    raw_gates = raw_config.get("gates", [])
    if not isinstance(raw_gates, list) or not raw_gates:
        raise ConfigurationError("Configuration must contain at least one [[gates]] entry.")

    gates: Dict[int, GateConfig] = {}
    for gate in raw_gates:
        try:
            gate_id = int(gate["id"])
            phone = normalize_phone_number(gate["phone"])
            password = str(gate["password"]).strip()
            sheet = str(gate["sheet"]).strip()
        except (KeyError, ValueError, TypeError) as exc:
            raise ConfigurationError(f"Invalid gate configuration: {gate!r}") from exc

        if not password:
            raise ConfigurationError(f"Gate {gate_id} is missing password.")
        if not sheet:
            raise ConfigurationError(f"Gate {gate_id} is missing sheet name.")
        if gate_id in gates:
            raise ConfigurationError(f"Gate {gate_id} is defined more than once.")

        gates[gate_id] = GateConfig(id=gate_id, phone=phone, password=password, sheet=sheet)

    return AppConfig(gateway_base=gateway_base, excel_path=excel_path, gates=gates)


class SmsGatewayClient:
    def __init__(self, config: AppConfig, session: requests.Session | None = None) -> None:
        self.config = config
        self.session = session or requests.Session()

    def send_sms(self, phone: str, message: str, timeout: float = DEFAULT_SEND_TIMEOUT) -> SmsResult:
        payload = {"phone": phone, "message": message}

        try:
            response = self.session.post(self.config.send_endpoint, json=payload, timeout=timeout)
            response.raise_for_status()
        except requests.exceptions.ConnectTimeout:
            error_msg = f"Connection to gateway {self.config.gateway_base} timed out."
            logger.error(error_msg)
            return SmsResult(ok=False, phone=phone, message=message, error=error_msg)
        except requests.exceptions.ConnectionError:
            error_msg = f"Failed to connect to gateway {self.config.gateway_base}."
            logger.error(error_msg)
            return SmsResult(ok=False, phone=phone, message=message, error=error_msg)
        except requests.exceptions.HTTPError:
            error_msg = f"Gateway returned HTTP {response.status_code}: {response.text}"
            logger.error(error_msg)
            return SmsResult(
                ok=False,
                phone=phone,
                message=message,
                error=error_msg,
                status_code=response.status_code,
                payload=response.text,
            )
        except requests.RequestException as exc:
            error_msg = f"Unexpected gateway error: {exc}"
            logger.error(error_msg)
            return SmsResult(ok=False, phone=phone, message=message, error=error_msg)

        try:
            payload_data = response.json()
        except ValueError:
            payload_data = response.text

        logger.info("SMS sent to %s", phone)
        return SmsResult(
            ok=True,
            phone=phone,
            message=message,
            status_code=response.status_code,
            payload=payload_data,
        )


def chunked(iterable: Iterable[str], size: int) -> Iterator[list[str]]:
    if size <= 0:
        raise ValueError("Chunk size must be greater than zero.")

    chunk: list[str] = []
    for item in iterable:
        chunk.append(item)
        if len(chunk) == size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


def normalize_phone_number(raw_value: object) -> str:
    if raw_value is None:
        raise ValueError("Empty value is not a phone number.")

    if pd.isna(raw_value):
        raise ValueError("Empty value is not a phone number.")

    if isinstance(raw_value, float) and raw_value.is_integer():
        text = str(int(raw_value))
    else:
        text = str(raw_value).strip()

    if not text:
        raise ValueError("Blank value is not a phone number.")

    compact = re.sub(r"[^\d+]", "", text)
    if compact.startswith("00"):
        compact = f"+{compact[2:]}"
    elif compact.startswith("+"):
        compact = f"+{re.sub(r'\\D', '', compact[1:])}"
    else:
        digits_only = re.sub(r"\D", "", compact)
        if len(digits_only) == 9:
            compact = f"+420{digits_only}"
        elif len(digits_only) == 12 and digits_only.startswith("420"):
            compact = f"+{digits_only}"
        else:
            raise ValueError(f"Unsupported phone number format: {raw_value!r}")

    if not PHONE_PATTERN.fullmatch(compact):
        raise ValueError(f"Unsupported phone number format: {raw_value!r}")

    return compact


def read_sheet(config: AppConfig, sheet_name: str) -> pd.DataFrame:
    try:
        with pd.ExcelFile(config.excel_path) as workbook:
            return pd.read_excel(workbook, sheet_name=sheet_name, header=None)
    except FileNotFoundError as exc:
        raise SpreadsheetError(f"Spreadsheet '{config.excel_path}' was not found.") from exc
    except ValueError as exc:
        raise SpreadsheetError(f"Sheet '{sheet_name}' could not be loaded: {exc}") from exc
    except OSError as exc:
        raise SpreadsheetError(f"Spreadsheet '{config.excel_path}' could not be opened: {exc}") from exc


def analyze_phone_numbers(values: Sequence[object], *, source_name: str) -> SheetAnalysis:
    rows: list[SheetRow] = []
    valid_numbers: list[str] = []
    unique_numbers: list[str] = []
    invalid_rows: list[SheetRow] = []
    blank_row_count = 0
    seen: set[str] = set()

    for row_number, value in enumerate(values, start=1):
        if pd.isna(value) or str(value).strip() == "":
            rows.append(SheetRow(row_number=row_number, raw_value="", normalized=None, status="blank"))
            blank_row_count += 1
            continue

        raw_value = str(value).strip()
        try:
            phone = normalize_phone_number(value)
        except ValueError as exc:
            row = SheetRow(
                row_number=row_number,
                raw_value=raw_value,
                normalized=None,
                status="invalid",
                error=str(exc),
            )
            rows.append(row)
            invalid_rows.append(row)
            continue

        status = "duplicate" if phone in seen else "valid"
        rows.append(
            SheetRow(
                row_number=row_number,
                raw_value=raw_value,
                normalized=phone,
                status=status,
            )
        )
        valid_numbers.append(phone)
        if phone not in seen:
            unique_numbers.append(phone)
            seen.add(phone)

    counts = Counter(valid_numbers)
    duplicate_numbers = [phone_number for phone_number, count in counts.items() if count > 1]

    return SheetAnalysis(
        source_name=source_name,
        rows=rows,
        valid_numbers=valid_numbers,
        unique_numbers=unique_numbers,
        duplicate_numbers=duplicate_numbers,
        invalid_rows=invalid_rows,
        blank_row_count=blank_row_count,
    )


def analyze_sheet(config: AppConfig, sheet_name: str, *, column_index: int) -> SheetAnalysis:
    dataframe = read_sheet(config, sheet_name)
    if dataframe.shape[1] <= column_index:
        raise SpreadsheetError(
            f"Sheet '{sheet_name}' does not contain column index {column_index + 1}."
        )

    values = dataframe.iloc[:, column_index].tolist()
    return analyze_phone_numbers(values, source_name=sheet_name)


def extract_phone_numbers(
    values: Sequence[object],
    *,
    source_name: str,
    deduplicate: bool = True,
) -> list[str]:
    analysis = analyze_phone_numbers(values, source_name=source_name)
    return normalize_phone_numbers_for_use(
        analysis,
        source_name=source_name,
        deduplicate=deduplicate,
        require_non_empty=True,
    )


def normalize_phone_numbers_for_use(
    analysis: SheetAnalysis,
    *,
    source_name: str,
    deduplicate: bool,
    require_non_empty: bool,
) -> list[str]:
    if analysis.invalid_rows:
        sample = "; ".join(
            f"{source_name} row {row.row_number}: {row.error}" for row in analysis.invalid_rows[:5]
        )
        suffix = "" if len(analysis.invalid_rows) <= 5 else f" (+{len(analysis.invalid_rows) - 5} more)"
        raise SpreadsheetError(f"Invalid phone numbers detected: {sample}{suffix}")

    numbers = analysis.unique_numbers if deduplicate else analysis.valid_numbers
    if require_non_empty and not numbers:
        raise SpreadsheetError(f"No phone numbers found in {source_name}.")

    return numbers


def get_sheet_numbers(
    config: AppConfig,
    sheet_name: str,
    *,
    column_index: int,
    deduplicate: bool = True,
) -> list[str]:
    analysis = analyze_sheet(config, sheet_name, column_index=column_index)
    return normalize_phone_numbers_for_use(
        analysis,
        source_name=sheet_name,
        deduplicate=deduplicate,
        require_non_empty=True,
    )


def prepare_numbers_for_sheet_save(values: Sequence[object], *, source_name: str) -> list[str]:
    analysis = analyze_phone_numbers(values, source_name=source_name)
    return normalize_phone_numbers_for_use(
        analysis,
        source_name=source_name,
        deduplicate=True,
        require_non_empty=False,
    )


def save_sheet_numbers(
    config: AppConfig,
    sheet_name: str,
    *,
    column_index: int,
    values: Sequence[object],
    create_backup: bool = False,
) -> SheetSaveResult:
    workbook_path = config.excel_path
    numbers = prepare_numbers_for_sheet_save(values, source_name=sheet_name)
    backup_path: Path | None = None

    if not workbook_path.exists():
        raise SpreadsheetError(f"Spreadsheet '{workbook_path}' was not found.")

    if create_backup:
        backup_path = create_backup_copy(workbook_path)

    try:
        workbook = load_workbook(BytesIO(workbook_path.read_bytes()))
    except FileNotFoundError as exc:
        raise SpreadsheetError(f"Spreadsheet '{workbook_path}' was not found.") from exc
    except OSError as exc:
        raise SpreadsheetError(f"Spreadsheet '{workbook_path}' could not be opened: {exc}") from exc

    try:
        worksheet = workbook[sheet_name]
    except KeyError as exc:
        raise SpreadsheetError(f"Sheet '{sheet_name}' was not found in '{workbook_path}'.") from exc

    target_column = column_index + 1
    max_rows_to_touch = max(worksheet.max_row or 0, len(numbers), 1)
    for row_index in range(1, max_rows_to_touch + 1):
        value = numbers[row_index - 1] if row_index <= len(numbers) else None
        worksheet.cell(row=row_index, column=target_column).value = value

    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=workbook_path.suffix,
            dir=workbook_path.parent,
        ) as temp_file:
            temp_path = Path(temp_file.name)

        workbook.save(temp_path)
        workbook.close()
        temp_path.replace(workbook_path)
    except PermissionError as exc:
        raise SpreadsheetError(
            f"Spreadsheet '{workbook_path}' could not be saved. It may be open in Excel."
        ) from exc
    except OSError as exc:
        raise SpreadsheetError(f"Spreadsheet '{workbook_path}' could not be saved: {exc}") from exc
    finally:
        workbook.close()
        del workbook
        gc.collect()
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)

    return SheetSaveResult(
        sheet_name=sheet_name,
        rows_written=len(numbers),
        numbers=numbers,
        backup_path=backup_path,
    )


def create_backup_copy(workbook_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = workbook_path.with_name(f"{workbook_path.stem}.bak_{timestamp}{workbook_path.suffix}")
    try:
        shutil.copy2(workbook_path, backup_path)
    except OSError as exc:
        raise SpreadsheetError(f"Backup for '{workbook_path}' could not be created: {exc}") from exc
    return backup_path


def build_add_messages(password: str, phone_numbers: Sequence[str], batch_size: int = DEFAULT_BATCH_SIZE) -> list[str]:
    return [f"{password} ADD {', '.join(block)}" for block in chunked(phone_numbers, batch_size)]


def build_find_messages(password: str, phone_numbers: Sequence[str]) -> list[str]:
    return [f"{password} FIND {phone_number}" for phone_number in phone_numbers]


def execute_messages(
    *,
    client: SmsGatewayClient,
    recipient_phone: str,
    messages: Sequence[str],
    dry_run: bool = False,
    pause_seconds: float = 0.0,
    stop_on_error: bool = True,
    timeout: float = DEFAULT_SEND_TIMEOUT,
    sleeper: Any = time.sleep,
) -> list[SmsResult]:
    results: list[SmsResult] = []

    for index, message in enumerate(messages, start=1):
        logger.info("Processing message %s/%s for %s", index, len(messages), recipient_phone)

        if dry_run:
            result = SmsResult(
                ok=True,
                phone=recipient_phone,
                message=message,
                payload={"dry_run": True},
            )
        else:
            result = client.send_sms(recipient_phone, message, timeout=timeout)

        results.append(result)

        if stop_on_error and not result.ok:
            break

        if not dry_run and pause_seconds > 0 and index < len(messages):
            sleeper(pause_seconds)

    return results


def poslat_sms(
    phone: str,
    message: str,
    *,
    timeout: float = DEFAULT_SEND_TIMEOUT,
    config: AppConfig | None = None,
    config_path: str | Path | None = None,
) -> Dict[str, Any]:
    active_config = config or load_config(config_path)
    client = SmsGatewayClient(active_config)
    normalized_phone = normalize_phone_number(phone)
    cleaned_message = message.strip()
    if not cleaned_message:
        raise ValueError("Message must not be empty.")

    result = client.send_sms(normalized_phone, cleaned_message, timeout=timeout)
    return {
        "ok": result.ok,
        "phone": result.phone,
        "message": result.message,
        "status_code": result.status_code,
        "payload": result.payload,
        "error": result.error,
    }


def poslat_davkove_sms(
    zvolena_zavora: int,
    *,
    batch_size: int = DEFAULT_BATCH_SIZE,
    pause_seconds: float = DEFAULT_ADD_DELAY_SECONDS,
    dry_run: bool = False,
    stop_on_error: bool = True,
    timeout: float = DEFAULT_SEND_TIMEOUT,
    config: AppConfig | None = None,
    config_path: str | Path | None = None,
) -> list[SmsResult]:
    active_config = config or load_config(config_path)
    gate = active_config.get_gate(zvolena_zavora)
    numbers = get_sheet_numbers(
        active_config,
        gate.sheet,
        column_index=GATE_PHONE_COLUMN_INDEX,
        deduplicate=True,
    )
    messages = build_add_messages(gate.password, numbers, batch_size=batch_size)
    client = SmsGatewayClient(active_config)
    return execute_messages(
        client=client,
        recipient_phone=gate.phone,
        messages=messages,
        dry_run=dry_run,
        pause_seconds=pause_seconds,
        stop_on_error=stop_on_error,
        timeout=timeout,
    )


def doplneni_seznamu_zavor(
    zvolena_zavora: int,
    *,
    batch_size: int = DEFAULT_BATCH_SIZE,
    pause_seconds: float = DEFAULT_ADD_DELAY_SECONDS,
    dry_run: bool = False,
    stop_on_error: bool = True,
    timeout: float = DEFAULT_SEND_TIMEOUT,
    config: AppConfig | None = None,
    config_path: str | Path | None = None,
) -> list[SmsResult]:
    active_config = config or load_config(config_path)
    gate = active_config.get_gate(zvolena_zavora)
    numbers = get_sheet_numbers(
        active_config,
        SUPPLEMENT_SHEET_NAME,
        column_index=SUPPLEMENT_PHONE_COLUMN_INDEX,
        deduplicate=True,
    )
    messages = build_add_messages(gate.password, numbers, batch_size=batch_size)
    client = SmsGatewayClient(active_config)
    return execute_messages(
        client=client,
        recipient_phone=gate.phone,
        messages=messages,
        dry_run=dry_run,
        pause_seconds=pause_seconds,
        stop_on_error=stop_on_error,
        timeout=timeout,
    )


def najit_cislo_na_zavore(
    zvolena_zavora: int,
    hledane_cislo: object,
    *,
    dry_run: bool = False,
    timeout: float = DEFAULT_SEND_TIMEOUT,
    config: AppConfig | None = None,
    config_path: str | Path | None = None,
) -> SmsResult:
    active_config = config or load_config(config_path)
    gate = active_config.get_gate(zvolena_zavora)
    searched_number = normalize_phone_number(hledane_cislo)
    message = f"{gate.password} FIND {searched_number}"
    client = SmsGatewayClient(active_config)
    return execute_messages(
        client=client,
        recipient_phone=gate.phone,
        messages=[message],
        dry_run=dry_run,
        pause_seconds=0.0,
        stop_on_error=True,
        timeout=timeout,
    )[0]


def najit_cisla_ze_seznamu_na_zavore(
    zvolena_zavora: int,
    *,
    pause_seconds: float = DEFAULT_FIND_DELAY_SECONDS,
    dry_run: bool = False,
    stop_on_error: bool = True,
    timeout: float = DEFAULT_SEND_TIMEOUT,
    config: AppConfig | None = None,
    config_path: str | Path | None = None,
) -> list[SmsResult]:
    active_config = config or load_config(config_path)
    gate = active_config.get_gate(zvolena_zavora)
    numbers = get_sheet_numbers(
        active_config,
        gate.sheet,
        column_index=GATE_PHONE_COLUMN_INDEX,
        deduplicate=True,
    )
    messages = build_find_messages(gate.password, numbers)
    client = SmsGatewayClient(active_config)
    return execute_messages(
        client=client,
        recipient_phone=gate.phone,
        messages=messages,
        dry_run=dry_run,
        pause_seconds=pause_seconds,
        stop_on_error=stop_on_error,
        timeout=timeout,
    )


def najit_duplikaty(
    zvolena_zavora: int,
    *,
    config: AppConfig | None = None,
    config_path: str | Path | None = None,
) -> list[str]:
    active_config = config or load_config(config_path)
    gate = active_config.get_gate(zvolena_zavora)
    numbers = get_sheet_numbers(
        active_config,
        gate.sheet,
        column_index=GATE_PHONE_COLUMN_INDEX,
        deduplicate=False,
    )
    counts = Counter(numbers)
    return [phone_number for phone_number, count in counts.items() if count > 1]
