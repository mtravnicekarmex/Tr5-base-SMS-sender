from __future__ import annotations

import gc
from io import BytesIO
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from send_sms import (
    SmsResult,
    analyze_sheet,
    analyze_phone_numbers,
    build_add_messages,
    chunked,
    execute_messages,
    extract_phone_numbers,
    load_config,
    normalize_phone_number,
    save_sheet_numbers,
)


class FakeClient:
    def __init__(self, results: list[SmsResult]) -> None:
        self._results = results
        self.calls: list[tuple[str, str, float]] = []

    def send_sms(self, phone: str, message: str, timeout: float = 10.0) -> SmsResult:
        self.calls.append((phone, message, timeout))
        return self._results[len(self.calls) - 1]


class SendSmsTests(unittest.TestCase):
    def test_chunked_splits_into_fixed_size_blocks(self) -> None:
        self.assertEqual(list(chunked(["1", "2", "3", "4", "5"], 2)), [["1", "2"], ["3", "4"], ["5"]])

    def test_normalize_phone_number_handles_local_and_float_values(self) -> None:
        self.assertEqual(normalize_phone_number("777 530 100"), "+420777530100")
        self.assertEqual(normalize_phone_number(777530100.0), "+420777530100")
        self.assertEqual(normalize_phone_number("+420 777 530 100"), "+420777530100")

    def test_extract_phone_numbers_deduplicates_and_skips_blank_rows(self) -> None:
        values = ["777530100", "777 530 100", None, "", "+420602191729"]
        self.assertEqual(
            extract_phone_numbers(values, source_name="sheet"),
            ["+420777530100", "+420602191729"],
        )

    def test_build_add_messages_groups_numbers(self) -> None:
        messages = build_add_messages("2803", ["+420111111111", "+420222222222", "+420333333333"], batch_size=2)
        self.assertEqual(
            messages,
            [
                "2803 ADD +420111111111, +420222222222",
                "2803 ADD +420333333333",
            ],
        )

    def test_analyze_phone_numbers_reports_duplicate_and_invalid_rows(self) -> None:
        analysis = analyze_phone_numbers(
            ["777530100", "777 530 100", "abc", None],
            source_name="sheet",
        )

        self.assertEqual(analysis.unique_numbers, ["+420777530100"])
        self.assertEqual(analysis.duplicate_numbers, ["+420777530100"])
        self.assertEqual(analysis.invalid_row_count, 1)
        self.assertEqual(analysis.blank_row_count, 1)

    def test_execute_messages_stops_after_first_error(self) -> None:
        results = [
            SmsResult(ok=True, phone="+420601060959", message="A"),
            SmsResult(ok=False, phone="+420601060959", message="B", error="gateway"),
            SmsResult(ok=True, phone="+420601060959", message="C"),
        ]
        client = FakeClient(results)

        executed = execute_messages(
            client=client,
            recipient_phone="+420601060959",
            messages=["A", "B", "C"],
            stop_on_error=True,
            pause_seconds=0.0,
        )

        self.assertEqual(len(executed), 2)
        self.assertEqual(len(client.calls), 2)

    def test_load_config_reads_gate_definitions(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            config_path = Path(temp_dir) / "config.toml"
            config_path.write_text(
                "\n".join(
                    [
                        'gateway_base = "http://localhost:8080"',
                        'excel_path = "data.xlsx"',
                        "",
                        "[[gates]]",
                        "id = 1",
                        'phone = "+420601060959"',
                        'password = "2803"',
                        'sheet = "1 - Benesovska"',
                    ]
                ),
                encoding="utf-8",
            )

            config = load_config(config_path)

            self.assertEqual(config.gateway_base, "http://localhost:8080")
            self.assertEqual(config.excel_path, Path("data.xlsx"))
            self.assertEqual(config.get_gate(1).phone, "+420601060959")

    def test_save_sheet_numbers_normalizes_and_preserves_other_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            workbook_path = temp_path / "gates.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Gate 1"
            worksheet.cell(row=1, column=1).value = "A1"
            worksheet.cell(row=1, column=2).value = "old1"
            worksheet.cell(row=1, column=3).value = "keep1"
            worksheet.cell(row=2, column=1).value = "A2"
            worksheet.cell(row=2, column=2).value = "old2"
            worksheet.cell(row=2, column=3).value = "keep2"
            worksheet.cell(row=3, column=1).value = "A3"
            worksheet.cell(row=3, column=2).value = "old3"
            worksheet.cell(row=3, column=3).value = "keep3"
            workbook.save(workbook_path)
            workbook.close()
            del workbook
            gc.collect()

            config_path = temp_path / "config.toml"
            config_path.write_text(
                "\n".join(
                    [
                        'gateway_base = "http://localhost:8080"',
                        f'excel_path = "{workbook_path.as_posix()}"',
                        "",
                        "[[gates]]",
                        "id = 1",
                        'phone = "+420601060959"',
                        'password = "2803"',
                        'sheet = "Gate 1"',
                    ]
                ),
                encoding="utf-8",
            )

            config = load_config(config_path)
            result = save_sheet_numbers(
                config,
                "Gate 1",
                column_index=1,
                values=["777530100", "+420777530100", "602191729", "", None],
            )

            self.assertEqual(result.rows_written, 2)
            self.assertEqual(result.numbers, ["+420777530100", "+420602191729"])

            saved_workbook = load_workbook(BytesIO(workbook_path.read_bytes()))
            saved_sheet = saved_workbook["Gate 1"]
            self.assertEqual(saved_sheet.cell(row=1, column=2).value, "+420777530100")
            self.assertEqual(saved_sheet.cell(row=2, column=2).value, "+420602191729")
            self.assertIsNone(saved_sheet.cell(row=3, column=2).value)
            self.assertEqual(saved_sheet.cell(row=1, column=1).value, "A1")
            self.assertEqual(saved_sheet.cell(row=2, column=3).value, "keep2")
            saved_workbook.close()
            del saved_workbook
            gc.collect()

            analysis = analyze_sheet(config, "Gate 1", column_index=1)
            self.assertEqual(analysis.unique_numbers, ["+420777530100", "+420602191729"])
            del analysis
            gc.collect()


if __name__ == "__main__":
    unittest.main()
